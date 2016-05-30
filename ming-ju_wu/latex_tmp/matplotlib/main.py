# -*- coding: utf-8 -*-

# encoding=utf8


__author__ = 'Michael'

"""
Bar chart demo with pairs of bars grouped for easy comparison.
"""


import numpy as np

#import matplotlib
import matplotlib as mpl
import matplotlib.ticker as plticker
from matplotlib.ticker import MultipleLocator, FormatStrFormatter
#mpl.use('Agg')

import matplotlib.pyplot as plt
from openpyxl import load_workbook

def example():
    n_groups = 5

    means_men = (20, 35, 30, 35, 27)
    std_men = (2, 3, 4, 1, 2)

    means_women = (25, 32, 34, 20, 25)
    std_women = (3, 5, 2, 3, 3)

    fig, ax = plt.subplots()

    index = np.arange(n_groups)
    bar_width = 0.35

    opacity = 0.4
    error_config = {'ecolor': '0.3'}

    rects1 = plt.bar(index, means_men, bar_width,
                     alpha=opacity,
                     color='b',
                     yerr=std_men,
                     error_kw=error_config,
                     label='Men')

    rects2 = plt.bar(index + bar_width, means_women, bar_width,
                     alpha=opacity,
                     color='r',
                     yerr=std_women,
                     error_kw=error_config,
                     label='Women')

    plt.xlabel('Group')
    plt.ylabel('Scores')
    plt.title('Scores by group and gender')
    plt.xticks(index + bar_width, ('A', 'B', 'C', 'D', 'E'))
    plt.legend()

    plt.tight_layout()
    plt.show()

def test2():
    dpoints = np.array([['rosetta', '1mfq', 9.97],
           ['rosetta', '1gid', 27.31],
           ['rosetta', '1y26', 5.77],
           ['rnacomposer', '1mfq', 5.55],
           ['rnacomposer', '1gid', 37.74],
           ['rnacomposer', '1y26', 5.77],
           ['random', '1mfq', 10.32],
           ['random', '1gid', 31.46],
           ['random', '1y26', 18.16]])


    fig = plt.figure()
    ax = fig.add_subplot(111)

    space = 0.3

    conditions = np.unique(dpoints[:,0])
    categories = np.unique(dpoints[:,1])

    n = len(conditions)

    width = (1 - space) / (len(conditions))
    print "width:", width

    for i,cond in enumerate(conditions):
        print "cond:", cond
        vals = dpoints[dpoints[:,0] == cond][:,2].astype(np.float)
        pos = [j - (1 - space) / 2. + i * width for j in range(1,len(categories)+1)]
        ax.bar(pos, vals, width=width)

    plt.show()

def plot_from_excel(sheetName, outputFileName):

    font = {'family' : 'Arial Unicode MS',
            'size'   : 13}
    mpl.rc('font', **font)

    wb = load_workbook('./Final Results-revise.xlsx', data_only=True)

    #ws = wb.worksheets[0]

    print (wb.get_sheet_names())
    ws = wb.get_sheet_by_name(sheetName)

    reference = []
    std_reference = []

    proposed = []
    std_proposed = []

    xticks = []

    if sheetName == u"工作表3":
        for i in range(0,4):
            for j in range(0,4):
                reference.append(ws.cell(row=(j*7 + i+3), column=11).value)
                std_reference.append(ws.cell(row=(j*7 + i+3), column=13).value)


                proposed.append(ws.cell(row=(j*7 + i+3), column=14).value)
                std_proposed.append(ws.cell(row=(j*7 + i+3), column=16).value)

        for i in range(0,16):
            xticks.append(ws.cell(row=29+i*2, column=0).value)

    else:
        for i in range(0,16):
            reference.append(ws.cell(row=2+i*2, column=2).value)
            std_reference.append(ws.cell(row=2+i*2, column=10).value)

            proposed.append(ws.cell(row=3+i*2, column=2).value)
            std_proposed.append(ws.cell(row=3+i*2, column=10).value)

            xticks.append(ws.cell(row=2+i*2, column=1).value)


    scale=2.0

    plt.rcParams['ytick.major.pad']='10'
    plt.rcParams['xtick.major.pad']='5'

    fig, ax = plt.subplots()
    fig.set_size_inches(6.3*scale,6.3/2*scale, forward=True)

    index = np.arange(16)
    bar_width = 0.275

    print index

    opacity = 1
    error_config = {'ecolor': '0.0', }

    print tuple(reference)
    print tuple(std_reference)

    ax.yaxis.grid(True)
    ax.set_axisbelow(True)

    rects1 = ax.bar(index+0.2, reference, bar_width,
                     #alpha=opacity,
                     color="#C0504D",
                     yerr=std_reference,
                     error_kw=error_config,
                     edgecolor = None,
                     linewidth=0,
                     label='Reference')

    print tuple(proposed)
    print tuple(std_proposed)

    rects2 = ax.bar(index + bar_width + 0.2 + 0.05, proposed, bar_width,
                     #alpha=opacity,
                     color="#1F497D",
                     yerr=std_proposed,
                     error_kw=error_config,
                     label='Proposed',
                     edgecolor = "w",
                     linewidth=0,
                     hatch="\\"*4
                    )



    box = ax.get_position()
    ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])
    if sheetName == u"工作表3":
        plt.subplots_adjust(left=0.09, right=0.85, top=0.95, bottom=0.12)
    else:
        plt.subplots_adjust(left=0.07, right=0.85, top=0.95, bottom=0.12)

    #plt.xlabel('Group')
    #plt.ylabel('MB/s')
    plt.ylabel(ws.cell(row=0, column=2).value)


    #plt.title('Scores by group and gender')

    plt.xticks(index + 0.5, xticks, rotation=45)

    for tic in ax.xaxis.get_major_ticks():
        tic.tick1On = False
        tic.tick2On = False

    for tic in ax.yaxis.get_major_ticks():
        tic.tick1On = False
        tic.tick2On = False

    # change tick resolution of yaxis
    if sheetName == u"工作表3":
        loc = plticker.MultipleLocator(base=2000) # this locator puts ticks at regular intervals
        ax.yaxis.set_major_locator(loc)



    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)


    #plt.legend()
    leg = ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
    leg.get_frame().set_linewidth(0.0)

    #plt.tight_layout()
    #plt.savefig("C:/Users/Michael/Desktop/tmp/plot.eps", format='eps')
    plt.savefig("./" + outputFileName + ".pdf")
    plt.show()

    print mpl.__version__

def plot_from_excel_2(sheetName, outputFileName):

    font = {'family' : 'Arial Unicode MS',
            'size'   : 13}
    mpl.rc('font', **font)

    wb = load_workbook(u"./實驗數據 - latency.xlsx", data_only=True)

    #ws = wb.worksheets[0]

    print (wb.get_sheet_names())
    ws = wb.get_sheet_by_name(sheetName)

    reference = []
    xticks = []

    for i in range(0,50):
        reference.append(ws.cell(row=2+i, column=2).value)

    for i in range(0,25):
        if sheetName == u"Latency - percent error - USED":
            xticks.append('%.2f' % (ws.cell(row=3+i*2, column=1).value))
        else:
            xticks.append(ws.cell(row=3+i*2, column=1).value)

    scale=2.0

    plt.rcParams['ytick.major.pad']='10'
    plt.rcParams['xtick.major.pad']='5'

    fig, ax = plt.subplots()
    fig.set_size_inches(6.3*scale,6.3/2*scale, forward=True)

    index = np.arange(50)
    bar_width = 0.5

    print index

    opacity = 1
    print tuple(reference)

    ax.yaxis.grid(True)
    ax.set_axisbelow(True)

    rects1 = ax.bar(index+0.25, reference, bar_width,
                     #alpha=opacity,
                     color="#C0504D",
                     #color="#1F497D",
                     edgecolor = None,
                     linewidth=0.5,
                    )

    ax.set_yscale('log')

    plt.xlabel(ws.cell(row=1, column=4).value)
    plt.ylabel(ws.cell(row=0, column=4).value)

    box = ax.get_position()
    if sheetName == u"Latency - percent error - USED":
        plt.subplots_adjust(left=0.07, right=0.97, top=0.93, bottom=0.15)
    else:
        plt.subplots_adjust(left=0.07, right=0.97, top=0.93, bottom=0.13)

    #plt.title('Scores by group and gender')

    #plt.xticks(index + 0.5, xticks)



    if sheetName == u"Latency - percent error - USED":
        plt.xticks(np.arange(1,50,2)+bar_width+0.125, xticks, rotation=45)
    else:
        plt.xticks(np.arange(1,50,2)+bar_width, xticks)

    #majorLocator   = MultipleLocator(5)
    minorLocator   = MultipleLocator(1)
    #ax.xaxis.set_major_locator(majorLocator)
    ax.xaxis.set_minor_locator(minorLocator)
    ax.tick_params(axis='x',which='minor',top='off')

    for tic in ax.xaxis.get_major_ticks():
        tic.tick1On = False
        tic.tick2On = False

    for tic in ax.yaxis.get_major_ticks():
        tic.tick1On = False
        tic.tick2On = False

    ax.spines['right'].set_visible(False)
    ax.spines['top'].set_visible(False)

    #ax.xaxis.set_ticks_position('bottom')

    #plt.tight_layout()
    #plt.savefig("C:/Users/Michael/Desktop/tmp/plot.eps", format='eps')
    #plt.savefig("./" + outputFileName + ".pdf")
    plt.show()

    #plt.savefig("./" + outputFileName + ".pdf")

if __name__ == "__main__":
    # Figures 9 ~ 12 of TECS journal paper
    #plot_from_excel(u"工作表1", "Figure 9")
    #plot_from_excel(u"工作表3", "Figure 10")
    #plot_from_excel(u"工作表4", "Figure 11")
    #plot_from_excel(u"工作表1 (3)", "Figure 12")

    # Figures 15(a) & 15(b) of StorageSim
    #plot_from_excel_2(u"Latency - USED", "Figure15a")
    plot_from_excel_2(u"Latency - percent error - USED", "Figure15b")

    #test2()